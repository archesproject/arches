"""
ARCHES - a program developed to inventory and manage immovable cultural heritage.
Copyright (C) 2013 J. Paul Getty Trust and World Monuments Fund

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as
published by the Free Software Foundation, either version 3 of the
License, or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program. If not, see <http://www.gnu.org/licenses/>.
"""

"""This module contains commands for building Arches."""

from django.core.management.base import BaseCommand
from openpyxl import Workbook, styles
from operator import itemgetter
from django.db import connection
from arches.app.models.card import Card
from arches.app.models.models import Node
from arches.app.models.system_settings import settings
from arches.app.utils.db_utils import dictfetchall


class Command(BaseCommand):
    """
    Commands creating data templates

    """

    def add_arguments(self, parser):
        parser.add_argument(
            "-t",
            "--template",
            action="store",
            dest="template",
            default="branchcsv",
            help="Type of template you would like to create - options include 'branchcsv'.",
        )
        parser.add_argument(
            "-d",
            "--dest",
            action="store",
            dest="dest",
            default="",
            help="Destination (directory and filename) where template should be saved. e.g. ~/Documents/mytemplate.xlsx",
        )
        parser.add_argument(
            "-g",
            "--graph",
            action="store",
            dest="graph",
            default="",
            help="Graphid for your template",
        )

    def handle(self, *args, **options):
        self.create_template(
            template=options["template"], dest=options["dest"], graphid=options["graph"]
        )

    def create_branchcsv_template(self, dest, graphid):
        wb = create_workbook(graphid)
        wb.save(filename=dest)

    def create_tilexls_template(self, dest, graphid):
        wb = create_tile_excel_workbook(graphid)
        wb.save(filename=dest)

    def create_template(self, template, dest, graphid):
        if template == "branchcsv":
            self.create_branchcsv_template(dest, graphid)
        elif template == "tilexls":
            self.create_tilexls_template(dest, graphid)


def write_metadata(workbook, metadata):
    metadata_sheet = workbook.create_sheet(title="metadata")
    metadata_sheet[f"A1"] = "graphid"
    metadata_sheet[f"B1"] = metadata["graphid"]
    metadata_sheet[f"A5"] = "node"
    metadata_sheet[f"B5"] = "datatype"
    for i, node in enumerate(metadata["nodes"]):
        metadata_sheet.cell(column=1, row=i + 6, value=node["alias"])
        metadata_sheet.cell(column=2, row=i + 6, value=node["datatype"])


def style_header(length, width, sheet):
    thin = styles.Side(border_style="thin", color="000000")
    fill = styles.PatternFill("solid", fgColor="eeeeee")
    border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in range(1, length + 1):
        for column in range(1, width + 1):
            sheet.cell(row=row, column=column).fill = fill
            sheet.cell(row=row, column=column).border = border


def create_workbook(graphid, tiledata=None) -> Workbook:
    wb = Workbook()
    columns = (
        "root_nodegroupid",
        "nodegroupid",
        "parent_nodegroupid",
        "alias",
        "name",
        "depth",
        "path",
        "cardinality",
    )
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT * FROM __get_nodegroup_tree_by_graph(%s)""", (graphid,)
        )
        rows = cursor.fetchall()
        first_sheet = True
        metadata = {"nodes": [], "graphid": graphid}
        column_length = 0
        for row in rows:
            details = dict(zip(columns, row))
            tab = "    " * details["depth"]
            nodes = (
                Node.objects.filter(nodegroup_id=details["nodegroupid"])
                .exclude(datatype="semantic")
                .values("datatype", "alias")
            )
            metadata["nodes"] += nodes
            if details["depth"] == 0:
                column_length = 0
                if first_sheet is True:
                    sheet = wb.active
                    sheet.title = details["alias"]
                    first_sheet = False
                else:
                    sheet = wb.create_sheet(title=details["alias"])
                row_number = 2
                sheet[f"A1"] = "resource_id"
                sheet[f"B1"] = "tileid"
                sheet[f"C1"] = "nodegroup"
                sheet[f"A{row_number}"] = "--"
                sheet[f"B{row_number}"] = "--"
                sheet[f"C2"] = f'{tab}{details["alias"]}  ({details["cardinality"]})'
                for i, node in enumerate(nodes):
                    sheet.cell(column=i + 4, row=row_number, value=f"{node['alias']}")
            else:
                row_number += 1
                sheet[f"A{row_number}"] = "--"
                sheet[f"B{row_number}"] = "--"
                sheet[f"C{row_number}"] = (
                    f'{tab}{details["alias"]}  ({details["cardinality"]})'
                )
                for i, node in enumerate(nodes):
                    sheet.cell(column=i + 4, row=row_number, value=f"{node['alias']}")
            column_length = len(nodes) if len(nodes) > column_length else column_length
            style_header(row_number, column_length + 3, sheet)

        if tiledata is not None:
            for root_nodegroup, tiles in tiledata.items():
                sheet = wb[root_nodegroup]
                for tile in tiles:
                    row_number = sheet.max_row + 1
                    tab = "    " * tile["depth"]
                    sheet[f"A{row_number}"] = str(tile["resourceinstanceid"])
                    sheet[f"B{row_number}"] = str(tile["tileid"])
                    sheet[f"C{row_number}"] = f'{tab}{tile["alias"]}'
                    nodes = (
                        Node.objects.filter(nodegroup_id=tile["nodegroupid"])
                        .exclude(datatype="semantic")
                        .values("alias")
                    )
                    for i, node in enumerate(nodes):
                        if tile[node["alias"]] is not None:
                            sheet.cell(
                                column=i + 4,
                                row=row_number,
                                value=f"{tile[node['alias']]}",
                            )

        write_metadata(wb, metadata)
        return wb


def create_tile_excel_workbook(graphid, tiledata=None):
    wb = Workbook()
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT * FROM __get_nodegroup_tree_by_graph(%s)""", (graphid,)
        )
        rows = dictfetchall(cursor)
        first_sheet = True
        for row in rows:
            card_name = str(Card.objects.get(nodegroup=row["nodegroupid"]).name)
            if first_sheet is True:
                sheet = wb.active
                sheet.title = card_name
                first_sheet = False
            else:
                sheet = wb.create_sheet(title=card_name)
            sheet[f"A1"] = "tileid"
            sheet[f"B1"] = "parenttile_id"
            sheet[f"C1"] = "resourceinstance_id"
            nodes = (
                Node.objects.filter(nodegroup_id=row["nodegroupid"])
                .exclude(datatype="semantic")
                .values("datatype", "alias")
            )
            for i, node in enumerate(nodes):
                sheet.cell(column=i + 4, row=1, value=node["alias"])
                sheet.cell(column=i + 5, row=1, value="sortorder")
                sheet.cell(column=i + 6, row=1, value="provisionaledits")
                sheet.cell(column=i + 7, row=1, value="nodegroup_id")

        if tiledata is not None:
            for card_name, tiles in tiledata.items():
                sheet = wb[card_name]
                for tile in tiles:
                    row_number = sheet.max_row + 1
                    sheet[f"A{row_number}"] = str(tile["tileid"])
                    sheet[f"B{row_number}"] = str(tile["parenttileid"])
                    sheet[f"C{row_number}"] = str(tile["resourceinstanceid"])
                    nodes = (
                        Node.objects.filter(nodegroup_id=tile["nodegroupid"])
                        .exclude(datatype="semantic")
                        .values("alias")
                    )
                    for i, node in enumerate(nodes):
                        if tile[node["alias"]] is not None:
                            sheet.cell(
                                column=i + 4,
                                row=row_number,
                                value=f"{tile[node['alias']]}",
                            )
                        sheet.cell(
                            column=i + 5, row=row_number, value=tile["sortorder"]
                        )
                        sheet.cell(
                            column=i + 6, row=row_number, value=tile["provisionaledits"]
                        )
                        sheet.cell(
                            column=i + 7, row=row_number, value=str(tile["nodegroupid"])
                        )

    return wb
