from datetime import datetime
import json
from openpyxl import load_workbook
import os
from tempfile import NamedTemporaryFile

from django.core.exceptions import ValidationError
import uuid
from django.db import connection
from django.http import HttpRequest, HttpResponse
from django.utils.translation import gettext as _
from django.core.files.storage import default_storage
from django.contrib.auth.models import User
from arches.app.datatypes.datatypes import DataTypeFactory
from arches.app.etl_modules.decorators import load_data_async
from arches.app.models.models import Node, TileModel, ETLModule
from arches.app.models.system_settings import settings
from arches.app.utils.betterJSONSerializer import JSONSerializer
from arches.app.etl_modules.base_import_module import (
    BaseImportModule,
    FileValidationError,
)
import arches.app.tasks as tasks
from arches.management.commands.etl_template import create_tile_excel_workbook


class TileExcelImporter(BaseImportModule):
    def __init__(self, request=None, loadid=None, temp_dir=None, params=None):
        self.loadid = request.POST.get("load_id") if request else loadid
        self.userid = (
            request.user.id
            if request
            else settings.DEFAULT_RESOURCE_IMPORT_USER["userid"]
        )
        self.mode = "cli" if not request and params else "ui"
        try:
            self.user = User.objects.get(pk=self.userid)
        except User.DoesNotExist:
            raise User.DoesNotExist(
                _(
                    "The userid {} does not exist. Probably DEFAULT_RESOURCE_IMPORT_USER is not configured correctly in settings.py.".format(
                        self.userid
                    )
                )
            )
        if not request and params:
            request = HttpRequest()
            request.user = self.user
            request.method = "POST"
            for k, v in params.items():
                request.POST.__setitem__(k, v)
        self.request = request
        self.moduleid = request.POST.get("module") if request else None
        self.datatype_factory = DataTypeFactory()
        self.legacyid_lookup = {}
        self.temp_path = ""
        self.temp_dir = temp_dir if temp_dir else None
        self.config = (
            ETLModule.objects.get(pk=self.moduleid).config if self.moduleid else {}
        )

    @load_data_async
    def run_load_task_async(self, request):
        self.loadid = request.POST.get("load_id")
        self.temp_dir = os.path.join(settings.UPLOADED_FILES_DIR, "tmp", self.loadid)
        self.file_details = request.POST.get("load_details", None)
        result = {}
        if self.file_details:
            details = json.loads(self.file_details)
            files = details["result"]["summary"]["files"]
            summary = details["result"]["summary"]

        load_task = tasks.load_tile_excel.apply_async(
            (self.userid, files, summary, result, self.temp_dir, self.loadid),
        )
        with connection.cursor() as cursor:
            cursor.execute(
                """UPDATE load_event SET taskid = %s WHERE loadid = %s""",
                (load_task.task_id, self.loadid),
            )

    def create_tile_value(
        self,
        cell_values,
        data_node_lookup,
        node_lookup,
        nodegroup_alias,
        row_details,
        cursor,
    ):
        node_value_keys = data_node_lookup[nodegroup_alias]
        tile_value = {}
        tile_valid = True
        for key in node_value_keys:
            try:
                nodeid = node_lookup[key]["nodeid"]
                node_details = node_lookup[key]
                datatype = node_details["datatype"]
                datatype_instance = self.datatype_factory.get_instance(datatype)
                source_value = row_details[key]
                config = node_details["config"]
                config["path"] = os.path.join(
                    settings.UPLOADED_FILES_DIR, "tmp", self.loadid
                )
                config["loadid"] = self.loadid
                try:
                    config["nodeid"] = nodeid
                except TypeError:
                    config = {}

                value, validation_errors = self.prepare_data_for_loading(
                    datatype_instance, source_value, config
                )
                valid = True if len(validation_errors) == 0 else False
                if not valid:
                    tile_valid = False
                error_message = ""
                for error in validation_errors:
                    error_message = (
                        "{0}|{1}".format(error_message, error["message"])
                        if error_message != ""
                        else error["message"]
                    )
                    cursor.execute(
                        """
                        INSERT INTO load_errors (type, value, source, error, message, datatype, loadid, nodeid)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (
                            "node",
                            source_value,
                            "",
                            error["title"],
                            error["message"],
                            datatype,
                            self.loadid,
                            nodeid,
                        ),
                    )

                tile_value[nodeid] = {
                    "value": value,
                    "valid": valid,
                    "source": source_value,
                    "notes": error_message,
                    "datatype": datatype,
                }
            except KeyError:
                pass

        tile_value_json = JSONSerializer().serialize(tile_value)
        return tile_value_json, tile_valid

    def process_worksheet(self, worksheet, cursor, node_lookup, nodegroup_lookup):
        data_node_lookup = {}
        row_count = 0

        nodegroupid_column = int(worksheet.max_column)
        maybe_nodegroup = worksheet.cell(row=2, column=nodegroupid_column).value
        if maybe_nodegroup:
            nodegroup_alias = nodegroup_lookup[maybe_nodegroup]["alias"]
            data_node_lookup[nodegroup_alias] = [
                val.value for val in worksheet[1][3:-3]
            ]
        # else: empty worksheet (no tiles)

        for row in worksheet.iter_rows(min_row=2):
            cell_values = [cell.value for cell in row]
            if len(cell_values) == 0 or any(cell_values) is False:
                continue
            resourceid = cell_values[2]
            if resourceid is None:
                cursor.execute(
                    """UPDATE load_event SET status = %s, load_end_time = %s WHERE loadid = %s""",
                    ("failed", datetime.now(), self.loadid),
                )
                raise ValueError(_("All rows must have a valid resource id"))

            node_values = cell_values[3:-3]
            try:
                row_count += 1
                row_details = dict(zip(data_node_lookup[nodegroup_alias], node_values))
                row_details["nodegroup_id"] = node_lookup[nodegroup_alias]["nodeid"]
                user_tileid = (
                    cell_values[0].strip()
                    if cell_values[0] and cell_values[0] != "None"
                    else None
                )
                tileid = user_tileid if user_tileid else uuid.uuid4()
                nodegroup_depth = nodegroup_lookup[row_details["nodegroup_id"]]["depth"]
                parenttileid = (
                    cell_values[1].strip()
                    if cell_values[1] and cell_values[1] != "None"
                    else None
                )
                legacyid, resourceid = self.set_legacy_id(resourceid)
                tile_value_json, passes_validation = self.create_tile_value(
                    cell_values,
                    data_node_lookup,
                    node_lookup,
                    nodegroup_alias,
                    row_details,
                    cursor,
                )
                nodegroup_cardinality = nodegroup_lookup[row_details["nodegroup_id"]][
                    "cardinality"
                ]
                operation = "insert"
                if user_tileid:
                    if nodegroup_cardinality == "n":
                        operation = (
                            "update"  # db will "insert" if tileid does not exist
                        )
                    elif nodegroup_cardinality == "1":
                        if TileModel.objects.filter(pk=tileid).exists():
                            operation = "update"
                cursor.execute(
                    """INSERT INTO load_staging (nodegroupid, legacyid, resourceid, tileid, parenttileid, value, loadid, nodegroup_depth, source_description, passes_validation, operation) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        row_details["nodegroup_id"],
                        legacyid,
                        resourceid,
                        tileid,
                        parenttileid,
                        tile_value_json,
                        self.loadid,
                        nodegroup_depth,
                        "worksheet:{0}, row:{1}".format(
                            worksheet.title, row[0].row
                        ),  # source_description
                        passes_validation,
                        operation,
                    ),
                )
            except KeyError:
                pass
        cursor.execute(
            """CALL __arches_check_tile_cardinality_violation_for_load(%s)""",
            [self.loadid],
        )
        cursor.execute(
            """
            INSERT INTO load_errors (type, source, error, loadid, nodegroupid)
            SELECT 'tile', source_description, error_message, loadid, nodegroupid
            FROM load_staging
            WHERE loadid = %s AND passes_validation = false AND error_message IS NOT null
            """,
            [self.loadid],
        )
        return {"name": worksheet.title, "rows": row_count}

    def validate_uploaded_file(self, workbook):
        graphid = None
        for worksheet in workbook.worksheets:
            if worksheet.cell(2, worksheet.max_column).value:
                try:
                    nodegroup_id = worksheet.cell(2, worksheet.max_column).value
                    graphid = str(
                        Node.objects.filter(nodegroup_id=nodegroup_id)[0].graph_id
                    )
                    break
                except (IndexError, ValidationError):
                    pass
        if graphid is None:
            raise FileValidationError()

    def get_graphid(self, workbook):
        for worksheet in workbook.worksheets:
            if worksheet.cell(2, worksheet.max_column).value:
                try:
                    nodegroup_id = worksheet.cell(2, worksheet.max_column).value
                    graphid = str(
                        Node.objects.filter(nodegroup_id=nodegroup_id)[0].graph_id
                    )
                    break
                except (IndexError, ValidationError):
                    pass
        return graphid

    def stage_files(self, files, summary, cursor):
        for file in files:
            self.stage_excel_file(file, summary, cursor)

    def stage_excel_file(self, file, summary, cursor):
        if file.endswith("xlsx"):
            summary["files"][file]["worksheets"] = []
            uploaded_file_path = os.path.join(
                settings.UPLOADED_FILES_DIR, "tmp", self.loadid, file
            )
            workbook = load_workbook(filename=default_storage.open(uploaded_file_path))
            graphid = self.get_graphid(workbook)
            nodegroup_lookup, nodes = self.get_graph_tree(graphid)
            node_lookup = self.get_node_lookup(nodes)

            for worksheet in workbook.worksheets:
                details = self.process_worksheet(
                    worksheet, cursor, node_lookup, nodegroup_lookup
                )
                summary["files"][file]["worksheets"].append(details)
            cursor.execute(
                """UPDATE load_event SET load_details = %s WHERE loadid = %s""",
                (json.dumps(summary), self.loadid),
            )

    def download(self, request):
        format = request.POST.get("format")
        if format == "xls":
            wb = create_tile_excel_workbook(request.POST.get("id"))
            with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                response = HttpResponse(
                    tmp.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = "attachment"
            os.unlink(tmp.name)
            return {"success": True, "raw": response}
        else:
            return {"success": False, "data": "failed"}
